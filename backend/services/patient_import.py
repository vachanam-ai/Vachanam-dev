"""Small, bounded parser for clinic patient migrations (CSV and XLSX)."""
from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from backend.services.validators import normalize_indian_phone
from backend.services.patient_identity import normalize_patient_age, normalize_patient_name

MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_EXPANDED_XLSX_BYTES = 50 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000

_ALIASES = {
    "name": {"name", "patient name", "full name", "patient", "patientname"},
    "phone": {
        "phone", "phone number", "mobile", "mobile number", "contact",
        "contact number", "contact no", "whatsapp", "patient mobile",
    },
    "age": {"age", "patient age", "years"},
    "gender": {"gender", "sex"},
}
_GENDERS = {
    "m": "male", "male": "male", "man": "male", "boy": "male",
    "f": "female", "female": "female", "woman": "female", "girl": "female",
    "o": "other", "other": "other", "nonbinary": "other", "non binary": "other",
}


class PatientImportError(ValueError):
    pass


@dataclass(frozen=True)
class ImportedPatient:
    row: int
    name: str
    phone: str
    age: int | None
    gender: str | None


def _header(value: object) -> str:
    text = re.sub(r"[_\-.]+", " ", str(value or "").strip().lower())
    return re.sub(r"\s+", " ", text)


def _columns(header: list[object]) -> dict[str, int]:
    normalized = [_header(value) for value in header]
    result: dict[str, int] = {}
    for field, aliases in _ALIASES.items():
        match = next((index for index, value in enumerate(normalized) if value in aliases), None)
        if match is not None:
            result[field] = match
    missing = [field for field in ("name", "phone") if field not in result]
    if missing:
        raise PatientImportError(
            "The first row must include Name and Mobile/Phone columns"
        )
    return result


def _cell(row: list[object], index: int | None) -> object:
    return row[index] if index is not None and index < len(row) else None


def _phone_text(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def _parse_rows(rows: Iterable[list[object]]) -> tuple[list[ImportedPatient], list[dict]]:
    iterator = iter(rows)
    try:
        columns = _columns(next(iterator))
    except StopIteration as exc:
        raise PatientImportError("The file is empty") from exc

    patients: list[ImportedPatient] = []
    errors: list[dict] = []
    for row_number, row in enumerate(iterator, start=2):
        if row_number > MAX_IMPORT_ROWS + 1:
            raise PatientImportError(f"A file can contain at most {MAX_IMPORT_ROWS:,} patients")
        if not any(str(value or "").strip() for value in row):
            continue
        try:
            name = normalize_patient_name(str(_cell(row, columns["name"]) or ""))
        except ValueError as exc:
            errors.append({"row": row_number, "error": str(exc)})
            continue
        try:
            phone = normalize_indian_phone(_phone_text(_cell(row, columns["phone"])))
        except ValueError:
            errors.append({"row": row_number, "error": "mobile number is invalid"})
            continue

        age: int | None = None
        raw_age = _cell(row, columns.get("age"))
        if raw_age not in (None, ""):
            try:
                age = normalize_patient_age(raw_age)
            except ValueError:
                errors.append({"row": row_number, "error": "age must be a whole number from 0 to 120"})
                continue

        raw_gender = _header(_cell(row, columns.get("gender")))
        gender = _GENDERS.get(raw_gender) if raw_gender else None
        patients.append(
            ImportedPatient(row=row_number, name=name, phone=phone, age=age, gender=gender)
        )
    return patients, errors


def parse_patient_file(filename: str, content: bytes) -> tuple[list[ImportedPatient], list[dict]]:
    """Parse only the four fields Vachanam uses; all other columns are ignored."""
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise PatientImportError("CSV must be saved as UTF-8") from exc
        return _parse_rows(csv.reader(io.StringIO(decoded)))

    if suffix == ".xlsx":
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                expanded = sum(item.file_size for item in archive.infolist())
                if expanded > MAX_EXPANDED_XLSX_BYTES:
                    raise PatientImportError("The Excel file expands beyond the 50 MB safety limit")
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except PatientImportError:
            raise
        except Exception as exc:
            raise PatientImportError("The Excel file is invalid or damaged") from exc
        try:
            sheet = workbook.active
            return _parse_rows((list(row) for row in sheet.iter_rows(values_only=True)))
        finally:
            workbook.close()

    raise PatientImportError("Upload a .csv or modern Excel .xlsx file")
